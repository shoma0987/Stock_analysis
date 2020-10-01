import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter import messagebox
import pandas as pd
import openpyxl as px
import tkinter .font as font

open_frame = tk.Tk()
open_frame.geometry("300x100")
open_frame.title("配当金管理アプリ")


label1 = tk.Label(open_frame, text="What's your username?")
label1.pack()


text = tk.Entry(open_frame)
text.pack()

wb = px.load_workbook("r_list.xlsx")
ws = wb["Sheet"]


def check_lines(): #名簿の確認
    i = 1
    for _ in range(100):
        if ws.cell(row= i,column=1).value != 0:
         i +=1
        else: break

    return "A" + str(i)

def registration_list():
    list1 = []
    for _ in range(100):
        if ws.cell(row=_+1,column=1).value != 0:
            nn = ws.cell(row=_+1,column=1).value
            list1.append(nn)
        else: break
    return list1

def scroll_bar2():
    with open("stock_225.txt","r",encoding="utf-8", errors="ignore") as f:
        r = f.read()
    N225 = list(r.split())

    a225,b225 =[],[]
    for i in range(0,450,2):
        a225.append(N225[i])
    for j in range(1,450,2):
        b225.append(N225[j])

    designated_window = tk.Tk()
    designated_window.title("N225")

    # スクロールバー
    scrollBar = Scrollbar(designated_window, orient=tk.VERTICAL)
    # スクロールバーのwidthは一番長い文字列に合わせる
    scrollBar.pack(side=RIGHT, fill=Y)

    myList = Listbox(designated_window, height=20, width=30, yscrollcommand=scrollBar.set)
    for line in range(225):
        myList.insert(END, a225[line] + " : " + b225[line])
    myList.pack()

    def add_stock():

        for i in myList.curselection():
            chosen = myList.get(i)
            designated_window.destroy()

            new_stocks = tk.Tk()
            new_stocks.title(chosen)
            new_stocks.geometry("300x150")

            label4 = tk.Label(new_stocks, text="Please input")

            label5 = tk.Label(new_stocks, text="Numbers of share")
            label5.pack()
            text2 = tk.Entry(new_stocks)
            text2.pack()

            label6 = tk.Label(new_stocks, text="Purchase price(¥) / per share")
            label6.pack()
            text3 = tk.Entry(new_stocks)
            text3.pack()

            okButton3 = ttk.Button(new_stocks, text="Confirm",command=lambda:[new_stocks.destroy()])
            okButton3.pack()

            new_stocks.mainloop()

    okButton2 = ttk.Button(designated_window, text="OK",command=lambda:[add_stock()])
    okButton2.pack()

    scrollBar.config(command=myList.yview)

def ok_check():
    s = text.get()
    open_frame.destroy() # OK押されたら、最初のwindowを閉じる
    if s in registration_list():
        # messagebox.showinfo("Welcome back, " + s + "!")
        Divi_page = tk.Tk()
        Divi_page.title("My page"+" (Welcome back, "+s+"!)")
        Divi_page.geometry("300x300")

        label2 = tk.Label(Divi_page, text="My Porfolio")
        label2.pack()


        label3 = tk.Label(Divi_page, text="Add new stocks!")
        label3.pack()
        Button3 = ttk.Button(Divi_page, text="Choose",command=lambda :[scroll_bar2()])
        Button3.pack()

        Divi_page.mainloop()

    else:
        messagebox.showinfo("Your username was registered now."+"["+ s +"]" )
        nl = check_lines()
        ws[nl] = s
        wb.save("r_list.xlsx")

#OK押した後に、強制的に最初のwindowを閉じる設定にする
okButton = ttk.Button(open_frame, text="OK", command=ok_check)
okButton.pack()

open_frame.mainloop()



# usernameが入力された新しいwindowとしてこちらのuserpageを表示させる
